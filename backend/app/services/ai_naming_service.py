"""
AI Naming Service using Google Gemini API for contextual Excel cell naming
"""
import os
import base64
import io
import logging
from typing import Dict, List, Optional, Tuple, Any
from pathlib import Path
import pandas as pd
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from PIL import Image
import google.generativeai as genai
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string

logger = logging.getLogger(__name__)

class AINameResult:
    """Result of AI naming for a single cell"""
    def __init__(self, cell_reference: str, suggested_name: str = None, 
                 confidence: float = 0.0, status: str = "failed", error_message: str = None):
        self.cell_reference = cell_reference
        self.suggested_name = suggested_name
        self.confidence = confidence
        self.status = status  # "success" or "failed"
        self.error_message = error_message

class AIBatchResult:
    """Result of batch AI naming for multiple cells"""
    def __init__(self):
        self.results: Dict[str, AINameResult] = {}
        self.failed_cells: List[str] = []
        self.processing_stats: Dict[str, Any] = {}

class AIExcelScreenshotGenerator:
    """Generate Excel-like screenshots for AI analysis"""
    
    def __init__(self, workbook_path: Path):
        self.workbook_path = workbook_path
        
    def generate_context_screenshot(self, sheet_name: str, target_cells: List[str], 
                                  context_rows: int = 12, use_extended_context: bool = True) -> bytes:
        """Generate a screenshot showing focused columns: A-E plus target cell columns only"""
        try:
            # Debug: Log the sheet name being requested
            logger.info(f"Generating screenshot for sheet: '{sheet_name}'")
            
            # Read Excel data with explicit sheet validation
            try:
                df = pd.read_excel(self.workbook_path, sheet_name=sheet_name, header=None)
                logger.info(f"Successfully loaded sheet '{sheet_name}' with {len(df)} rows and {len(df.columns)} columns")
            except ValueError as e:
                # Sheet name doesn't exist, log available sheets
                available_sheets = pd.ExcelFile(self.workbook_path).sheet_names
                logger.error(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")
                raise ValueError(f"Sheet '{sheet_name}' not found in workbook. Available sheets: {available_sheets}")
            except Exception as e:
                logger.error(f"Error reading Excel file for sheet '{sheet_name}': {e}")
                raise
            
            # Get row range and target columns from target cells
            min_row, max_row, target_columns = self._get_focused_range(target_cells)
            
            if use_extended_context:
                # Extended context: Focused columns (A-E + target columns), row 1 to target row
                start_row = 0  # Always start from row 1
                end_row = max_row + 1  # Include up to the target cell row
                
                # Select focused columns: A-E (0-4) + unique target columns
                context_columns = [0, 1, 2, 3, 4]  # A-E
                selected_columns = context_columns + [col for col in target_columns if col not in context_columns]
                
                # Ensure we don't exceed DataFrame bounds
                selected_columns = [col for col in selected_columns if col < len(df.columns)]
                selected_columns.sort()  # Keep in order for readability
                
                logger.info(f"Using focused columns: {[self._get_column_letter(col + 1) for col in selected_columns]}")
                
            else:
                # Legacy context: limited rows and columns A-E only
                start_row = max(0, min_row - context_rows // 2)
                end_row = min(len(df), max_row + context_rows // 2)
                selected_columns = list(range(min(5, len(df.columns))))  # Columns A-E (indices 0-4)
            
            # Extract subset using selected columns
            subset_df = df.iloc[start_row:end_row, selected_columns].copy()
            
            # Create figure
            fig, ax = plt.subplots(figsize=(14, 10))
            ax.axis('tight')
            ax.axis('off')
            
            # Create table
            table_data = []
            for i, row in subset_df.iterrows():
                table_data.append([str(val) if pd.notna(val) else '' for val in row])
            
            # Add column headers based on actual selected columns
            if use_extended_context:
                col_headers = [self._get_column_letter(col + 1) for col in selected_columns]
            else:
                col_headers = [get_column_letter(i + 1) for i in range(len(subset_df.columns))]
            
            # Add row numbers
            row_headers = [str(start_row + i + 1) for i in range(len(subset_df))]
            
            # Create table with headers
            table = ax.table(cellText=table_data, 
                           colLabels=col_headers,
                           rowLabels=row_headers,
                           cellLoc='left',
                           loc='center',
                           colWidths=[0.12] * len(col_headers))
            
            # Style the table
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1, 1.5)
            
            # Add title
            if use_extended_context:
                col_range = f'A-E + {", ".join([self._get_column_letter(col + 1) for col in target_columns])}'
                plt.title(f'Sheet: {sheet_name} - Focused Context ({col_range}) for AI Analysis', 
                         fontsize=14, fontweight='bold', pad=20)
            else:
                plt.title(f'Sheet: {sheet_name} - Context Columns A-E for AI Analysis', 
                         fontsize=14, fontweight='bold', pad=20)
            
            # Save to bytes
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', dpi=150, bbox_inches='tight', 
                       facecolor='white', edgecolor='none')
            buffer.seek(0)
            screenshot_bytes = buffer.getvalue()
            plt.close(fig)
            
            return screenshot_bytes
            
        except Exception as e:
            logger.error(f"Error generating screenshot for {sheet_name}: {e}")
            # Return empty screenshot on error
            fig, ax = plt.subplots(figsize=(10, 6))
            ax.text(0.5, 0.5, f'Error generating screenshot:\n{str(e)}', 
                   ha='center', va='center', fontsize=12)
            ax.axis('off')
            buffer = io.BytesIO()
            plt.savefig(buffer, format='png', bbox_inches='tight')
            buffer.seek(0)
            screenshot_bytes = buffer.getvalue()
            plt.close(fig)
            return screenshot_bytes
    
    def _get_focused_range(self, cell_addresses: List[str]) -> Tuple[int, int, List[int]]:
        """Get focused range (min_row, max_row, unique_target_columns) for cell addresses"""
        rows, cols = [], []
        
        for cell_addr in cell_addresses:
            try:
                col_letter, row_num = self._parse_cell_address(cell_addr)
                col_idx = column_index_from_string(col_letter) - 1  # 0-based
                row_idx = row_num - 1  # 0-based
                rows.append(row_idx)
                cols.append(col_idx)
            except:
                continue
        
        if not rows or not cols:
            return 0, 10, [0, 1, 2, 3, 4]  # Default: A-E columns
        
        # Return unique target columns (sorted, no duplicates)
        unique_target_cols = sorted(list(set(cols)))
        return min(rows), max(rows), unique_target_cols
    
    def _parse_cell_address(self, cell_address: str) -> Tuple[str, int]:
        """Parse cell address into column letter and row number"""
        import re
        if '!' in cell_address:
            cell_address = cell_address.split('!')[1]
        
        match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
        if not match:
            raise ValueError(f"Invalid cell address: {cell_address}")
        
        return match.groups()[0], int(match.groups()[1])
    
    def _get_column_letter(self, col_num: int) -> str:
        """Convert column number to letter (1=A, 2=B, etc.)"""
        from openpyxl.utils import get_column_letter
        return get_column_letter(col_num)

class AINameService:
    """Service for AI-powered contextual cell naming using Gemini"""
    
    def __init__(self):
        self.api_key = os.getenv('GEMINI_API_KEY')
        if self.api_key:
            genai.configure(api_key=self.api_key)
            self.model = genai.GenerativeModel('gemini-2.5-flash')
        else:
            logger.warning("GEMINI_API_KEY not found. AI naming will be disabled.")
            self.model = None
    
    def group_cells_by_sheet(self, cell_references: List[str]) -> Dict[str, List[str]]:
        """Group cell references by sheet name"""
        sheet_groups = {}
        
        for cell_ref in cell_references:
            if '!' in cell_ref:
                sheet_name, cell_addr = cell_ref.split('!', 1)
                if sheet_name not in sheet_groups:
                    sheet_groups[sheet_name] = []
                sheet_groups[sheet_name].append(cell_ref)
            else:
                # Default sheet if no sheet specified (shouldn't happen in this app)
                if 'default' not in sheet_groups:
                    sheet_groups['default'] = []
                sheet_groups['default'].append(cell_ref)
        
        logger.info(f"Grouped cells by sheet: {dict((k, len(v)) for k, v in sheet_groups.items())}")
        return sheet_groups
    
    async def generate_batch_names(self, workbook_path: Path, sheet_name: str, 
                                 cell_references: List[str], 
                                 structured_data: Dict = None, use_extended_context: bool = True) -> AIBatchResult:
        """Generate AI names for a batch of cells using sheet-based processing"""
        if not self.model:
            # API key not configured
            result = AIBatchResult()
            for cell_ref in cell_references:
                result.results[cell_ref] = AINameResult(
                    cell_reference=cell_ref,
                    status="failed",
                    error_message="Gemini API not configured. Set GEMINI_API_KEY environment variable."
                )
                result.failed_cells.append(cell_ref)
            return result
        
        # Group cells by sheet for proper processing
        sheet_groups = self.group_cells_by_sheet(cell_references)
        
        # Process each sheet separately
        return await self.generate_batch_names_by_sheet(workbook_path, sheet_groups, use_extended_context)
    
    async def generate_batch_names_by_sheet(self, workbook_path: Path, 
                                          sheet_groups: Dict[str, List[str]], use_extended_context: bool = True) -> AIBatchResult:
        """Generate AI names by processing each sheet separately"""
        combined_result = AIBatchResult()
        
        for sheet_name, cell_refs in sheet_groups.items():
            logger.info(f"Processing {len(cell_refs)} cells from sheet '{sheet_name}'")
            
            try:
                # Generate screenshot for THIS sheet
                screenshot_gen = AIExcelScreenshotGenerator(workbook_path)
                screenshot_bytes = screenshot_gen.generate_context_screenshot(sheet_name, cell_refs, use_extended_context=use_extended_context)
                
                # Create sheet-specific prompt with line numbers
                prompt = self._create_sheet_specific_prompt(sheet_name, cell_refs, use_extended_context)
                
                # Process this sheet's cells
                sheet_result = await self._process_single_sheet(prompt, screenshot_bytes, cell_refs)
                
                # Merge results
                combined_result.results.update(sheet_result.results)
                combined_result.failed_cells.extend(sheet_result.failed_cells)
                
            except Exception as e:
                logger.error(f"Error processing sheet '{sheet_name}': {e}")
                # Mark all cells in this sheet as failed
                for cell_ref in cell_refs:
                    combined_result.results[cell_ref] = AINameResult(
                        cell_reference=cell_ref,
                        status="failed",
                        error_message=f"Sheet processing failed: {str(e)}"
                    )
                    combined_result.failed_cells.append(cell_ref)
        
        # Update combined statistics
        total_cells = sum(len(cells) for cells in sheet_groups.values())
        successful = len([r for r in combined_result.results.values() if r.status == "success"])
        combined_result.processing_stats = {
            "total_cells": total_cells,
            "successful": successful,
            "failed": total_cells - successful,
            "sheets_processed": len(sheet_groups)
        }
        
        return combined_result
    
    async def _process_single_sheet(self, prompt: str, screenshot_bytes: bytes, 
                                  cell_references: List[str]) -> AIBatchResult:
        """Process a single sheet's cells with AI"""
        result = AIBatchResult()
        
        try:
            # Convert screenshot to base64 for Gemini
            screenshot_base64 = base64.b64encode(screenshot_bytes).decode('utf-8')
            
            # Create image part for Gemini
            image_part = {
                "mime_type": "image/png",
                "data": screenshot_base64
            }
            
            # Call Gemini API
            response = self.model.generate_content([prompt, image_part])
            
            # Parse response
            if response and response.text:
                parsed_results = self._parse_gemini_response(response.text, cell_references)
                result.results.update(parsed_results)
            else:
                raise Exception("No response from Gemini API")
                
        except Exception as e:
            logger.error(f"Error in single sheet AI processing: {e}")
            # Mark all cells as failed
            for cell_ref in cell_references:
                result.results[cell_ref] = AINameResult(
                    cell_reference=cell_ref,
                    status="failed",
                    error_message=f"AI generation failed: {str(e)}"
                )
                result.failed_cells.append(cell_ref)
        
        return result
    
    def _create_sheet_specific_prompt(self, sheet_name: str, cell_references: List[str], use_extended_context: bool = True) -> str:
        """Create sheet-specific prompt focused on relevant line numbers"""
        
        # Extract row numbers for cells that belong to THIS sheet only
        sheet_rows = []
        for cell_ref in cell_references:
            try:
                if '!' in cell_ref:
                    ref_sheet, cell_addr = cell_ref.split('!', 1)
                    # Only process cells that belong to this sheet
                    if ref_sheet == sheet_name:
                        import re
                        match = re.match(r'^[A-Z]+(\d+)$', cell_addr.upper())
                        if match:
                            row_num = int(match.group(1))
                            sheet_rows.append((cell_ref, row_num))
            except Exception as e:
                logger.error(f"Error parsing cell reference {cell_ref}: {e}")
                continue

        # Create line-specific instructions for this sheet
        line_instructions = []
        for cell_ref, row_num in sheet_rows:
            line_instructions.append(f'- {cell_ref} (row {row_num})')
        
        line_list = '\n'.join(line_instructions) if line_instructions else "- No valid cells found for this sheet"
        
        if use_extended_context:
            prompt = f"""You are analyzing the '{sheet_name}' worksheet of a financial Excel model. 
The screenshot shows focused columns: A-E (for context) plus the specific target cell columns, and all rows from 1 to the target cell rows.

Your task is to create period-aware descriptive names for the following cells using a two-step lookup process:

{line_list}

NAMING PROCESS - Follow these steps for EACH cell:

Step 1 - PERIOD EXTRACTION:
- For PERIOD information: Look ONLY at the first 10 rows of the specific column containing the target cell
- Example: For a target cell in column X row Y, examine only column X rows 1-10 for period indicators
- Ignore period information from other columns - focus solely on the target cell's column header area
- Look for: years (2024, 2025), quarters (Q1, Q2, 1Q, 2Q), months (Jan, Feb), fiscal years (FY24), etc.

Step 2 - DESCRIPTION EXTRACTION:
- For DESCRIPTION information: Look at the first 5 columns (A-E typically) of the target cell's row for descriptive text
- Example: For a target cell in column X row Y, examine columns A-E of row Y for business context
- Use this row context to determine what the line item represents
- Look for: Revenue, EBITDA, Operating Income, Expenses, etc.

Step 3 - COMBINE:
- Combine as "[Period] [Description]" using the most specific period available
- Example: "2025 Revenue", "Q1 Operating Expenses", "Jan 2024 EBITDA"
- If no clear period is found in the target column, use description only

Requirements:
- Maximum 75 characters per name
- Use the most specific period available (e.g., "2Q25" instead of just "Q2" or "2025")
- Follow financial modeling conventions
- Process each cell individually using the two-step lookup above

CRITICAL - NO HALLUCINATION: Only use information that is explicitly visible in the screenshot.
Do NOT make assumptions about what the data should contain based on context clues or common business patterns.
If you cannot clearly see descriptive text in the row's first 5 columns, use generic terms like "Line Item", "Revenue Item", or "Segment Item".
Base your names ONLY on text that is actually readable in the screenshot - never infer or assume content.

Return your response as a JSON object with a single key, "cell_names", where the value is another object mapping cell references to their suggested name and confidence score.
For example:
{{
  "cell_names": {{
    "Sheet1!A1": {{'name': "2025 Revenue", 'confidence': 0.95}},
    "Sheet1!B5": {{'name': "Q1 Operating Expenses", 'confidence': 0.80}},
    "Sheet1!C10": {{'name': "Line Item", 'confidence': 0.60}}
  }}
}}
"""
        else:
            # Legacy prompt (original behavior)
            prompt = f"""You are analyzing the '{sheet_name}' worksheet of a financial Excel model. 
The screenshot shows columns A-E of this sheet.

Your task is to create descriptive names for the following cells by analyzing their row context in the screenshot:

{line_list}

Requirements:
- Maximum 50 characters per name.
- Include business context (e.g., Revenue, EBITDA, Operating Income).
- Be specific about the line item type.
- Follow financial modeling conventions.
- Do NOT include years, quarters, or time periods in names.
- Focus only on describing WHAT the line item represents, not WHEN.
- Note: Only columns A-E are visible - year information will be added separately.

Return your response as a JSON object with a single key, "cell_names", where the value is another object mapping cell references to their suggested name and confidence score.
For example:
{{
  "cell_names": {{
    "Sheet1!A1": {{'name': "Descriptive Name", 'confidence': 0.95}}
  }}
}}
"""
        return prompt
    
    def _parse_gemini_response(self, response_text: str, cell_references: List[str]) -> Dict[str, AINameResult]:
        """Parse Gemini API response and create AINameResult objects"""
        results = {}
        
        try:
            import json
            
            # Try to extract JSON from response
            response_text = response_text.strip()
            if response_text.startswith('```json'):
                response_text = response_text[7:]
            if response_text.endswith('```'):
                response_text = response_text[:-3]
            
            data = json.loads(response_text)
            cell_names = data.get('cell_names', {})
            
            for cell_ref in cell_references:
                if cell_ref in cell_names:
                    cell_data = cell_names[cell_ref]
                    results[cell_ref] = AINameResult(
                        cell_reference=cell_ref,
                        suggested_name=cell_data.get('name', ''),
                        confidence=cell_data.get('confidence', 0.0),
                        status="success"
                    )
                else:
                    results[cell_ref] = AINameResult(
                        cell_reference=cell_ref,
                        status="failed",
                        error_message="AI could not generate name for this cell"
                    )
                    
        except Exception as e:
            logger.error(f"Error parsing Gemini response: {e}")
            # Fall back to failed status for all cells
            for cell_ref in cell_references:
                results[cell_ref] = AINameResult(
                    cell_reference=cell_ref,
                    status="failed",
                    error_message=f"Failed to parse AI response: {str(e)}"
                )
        
        return results
    
    async def generate_table_summary(self, baseline_data: List[Dict], prompt: str) -> Dict[str, Any]:
        """Generate AI summary for table data using screenshot analysis"""
        try:
            if not self.model:
                return {
                    "status": "failed",
                    "error_message": "Gemini API not configured. Set GEMINI_API_KEY environment variable."
                }
            
            # Generate table screenshot
            screenshot_bytes = self._generate_table_screenshot(baseline_data)
            
            # Create image for Gemini
            import PIL.Image
            image = PIL.Image.open(io.BytesIO(screenshot_bytes))
            
            # Send to Gemini with prompt
            response = await self.model.generate_content_async([
                prompt,
                image
            ])
            
            if response.text:
                return {
                    "status": "success",
                    "summary": response.text.strip()
                }
            else:
                return {
                    "status": "failed",
                    "error_message": "Gemini returned empty response"
                }
                
        except Exception as e:
            logger.error(f"Error generating table summary: {e}")
            return {
                "status": "failed",
                "error_message": str(e)
            }
    
    def _generate_table_screenshot(self, baseline_data: List[Dict]) -> bytes:
        """Generate screenshot of the baseline table with proper formula formatting"""
        try:
            # Create a figure for the table
            fig, ax = plt.subplots(figsize=(14, max(6, len(baseline_data) * 0.4 + 2)))
            ax.axis('tight')
            ax.axis('off')
            
            # Prepare table data with headers
            headers = ['CELL REFERENCE', 'NAME', 'VALUE', 'FORMULA']
            table_data = [headers]
            
            # Add data rows with proper formula formatting
            for row in baseline_data:
                formatted_row = [
                    str(row.get('cellReference', '')),
                    str(row.get('name', '')),
                    str(row.get('value', '')),
                    f'"{str(row.get("formula", ""))}"' if row.get('formula') else '-'  # Wrap formulas in quotes
                ]
                table_data.append(formatted_row)
            
            # Create table with styling
            table = ax.table(
                cellText=table_data[1:],  # Data rows
                colLabels=table_data[0],  # Headers
                cellLoc='left',
                loc='center',
                colWidths=[0.25, 0.3, 0.15, 0.3]  # Adjust column widths
            )
            
            # Style the table
            table.auto_set_font_size(False)
            table.set_fontsize(9)
            table.scale(1.2, 1.8)
            
            # Header styling
            for i in range(len(headers)):
                cell = table[(0, i)]
                cell.set_facecolor('#4472C4')
                cell.set_text_props(weight='bold', color='white')
                
            # Row styling based on type
            for i, row in enumerate(baseline_data, 1):
                row_type = row.get('rowType', 'constant')
                color = '#E7F3FF' if row_type == 'formula' else '#E8F5E8'  # Blue for formula, green for constant
                
                for j in range(len(headers)):
                    cell = table[(i, j)]
                    cell.set_facecolor(color)
            
            # Save to bytes
            buf = io.BytesIO()
            plt.savefig(buf, format='png', bbox_inches='tight', dpi=150, facecolor='white', edgecolor='none')
            plt.close()
            buf.seek(0)
            
            return buf.getvalue()
            
        except Exception as e:
            logger.error(f"Error generating table screenshot: {e}")
            raise
    
    async def generate_variance_summary(self, baseline_data: List[Dict], new_data: List[Dict], prompt: str) -> Dict[str, Any]:
        """Generate AI variance summary comparing two tables using dual screenshot analysis"""
        try:
            if not self.model:
                return {
                    "status": "failed",
                    "error_message": "Gemini API not configured. Set GEMINI_API_KEY environment variable."
                }
            
            # Generate screenshots for both tables
            baseline_screenshot = self._generate_table_screenshot(baseline_data)
            new_screenshot = self._generate_table_screenshot(new_data)
            
            # Create images for Gemini
            import PIL.Image
            baseline_image = PIL.Image.open(io.BytesIO(baseline_screenshot))
            new_image = PIL.Image.open(io.BytesIO(new_screenshot))
            
            # Send to Gemini with both images and comparison prompt
            response = await self.model.generate_content_async([
                prompt,
                "BASELINE Table:",
                baseline_image,
                "NEW Table:",
                new_image
            ])
            
            if response.text:
                return {
                    "status": "success",
                    "summary": response.text.strip()
                }
            else:
                return {
                    "status": "failed",
                    "error_message": "Gemini returned empty response"
                }
                
        except Exception as e:
            logger.error(f"Error generating variance summary: {e}")
            return {
                "status": "failed",
                "error_message": str(e)
            }
