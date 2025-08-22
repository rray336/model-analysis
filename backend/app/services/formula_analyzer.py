"""
Excel Formula Analysis Engine
Parses Excel formulas to build drill-down dependency trees
"""
import re
import openpyxl
from pathlib import Path
from typing import Dict, List, Set, Optional, Tuple, Any
from dataclasses import dataclass, field
import logging

logger = logging.getLogger(__name__)

@dataclass
class CellReference:
    """Represents a single cell reference"""
    sheet_name: Optional[str] = None
    column: str = ""
    row: int = 0
    is_absolute_column: bool = False
    is_absolute_row: bool = False
    is_range: bool = False
    range_end_column: Optional[str] = None
    range_end_row: Optional[int] = None
    
    def __str__(self):
        abs_col = "$" if self.is_absolute_column else ""
        abs_row = "$" if self.is_absolute_row else ""
        
        if self.is_range:
            range_abs_col = "$" if self.is_absolute_column else ""
            range_abs_row = "$" if self.is_absolute_row else ""
            base = f"{abs_col}{self.column}{abs_row}{self.row}"
            end = f"{range_abs_col}{self.range_end_column}{range_abs_row}{self.range_end_row}"
            result = f"{base}:{end}"
        else:
            result = f"{abs_col}{self.column}{abs_row}{self.row}"
        
        if self.sheet_name:
            return f"{self.sheet_name}!{result}"
        return result

@dataclass
class FormulaComponent:
    """Represents a component of a formula breakdown"""
    name: str
    cell_reference: str
    value: float
    formula: Optional[str] = None
    is_leaf_node: bool = False
    dependencies: List['FormulaComponent'] = field(default_factory=list)
    variance_contribution: float = 0.0

@dataclass
class DrillDownResult:
    """Result of drilling down into a formula"""
    source_item: str
    source_value: float
    components: List[FormulaComponent]
    total_explained: float
    unexplained_variance: float
    drill_down_path: List[str]

class FormulaAnalyzer:
    """Analyzes Excel formulas for drill-down capabilities"""
    
    def __init__(self):
        # Compile regex patterns for performance
        self.cell_pattern = re.compile(
            r"(?:(['\w\s]+)!)?"  # Optional sheet name with quotes
            r"(\$?[A-Z]+)(\$?\d+)"  # Column and row  
            r"(?::(\$?[A-Z]+)(\$?\d+))?"  # Optional range end
        )
        
        self.simple_cell_pattern = re.compile(r"([A-Za-z]+)(\d+)", re.IGNORECASE)
        
        # Function patterns we can analyze
        self.supported_functions = {
            'SUM', 'AVERAGE', 'COUNT', 'MAX', 'MIN', 'SUMIF', 'SUMIFS',
            'IF', 'IFERROR', 'VLOOKUP', 'HLOOKUP', 'INDEX', 'MATCH'
        }
        
    def parse_formula(self, formula: str) -> List[CellReference]:
        """Parse an Excel formula and extract cell references"""
        if not formula or not formula.startswith('='):
            return []
        
        # Clean up the formula
        formula_clean = formula[1:]  # Remove leading =
        
        # Check for external file references (e.g., [Workbook.xlsx]Sheet1!A1)
        if '[' in formula_clean and ']' in formula_clean:
            logger.info(f"External file reference detected in formula: {formula[:100]}...")
        
        references = []
        matches = self.cell_pattern.findall(formula_clean)
        
        for match in matches:
            sheet_name, col, row, range_col, range_row = match

            try:
                # Clean sheet name (remove quotes if present)
                if sheet_name:
                    sheet_name = sheet_name.strip("'\"")
                
                # Parse absolute references
                is_abs_col = col.startswith('$')
                is_abs_row = row.startswith('$')
                
                col_clean = col.replace('$', '')
                row_clean = int(row.replace('$', ''))
                
                ref = CellReference(
                    sheet_name=sheet_name if sheet_name else None,
                    column=col_clean,
                    row=row_clean,
                    is_absolute_column=is_abs_col,
                    is_absolute_row=is_abs_row
                )
                
                # Handle ranges
                if range_col and range_row:
                    ref.is_range = True
                    ref.range_end_column = range_col.replace('$', '')
                    ref.range_end_row = int(range_row.replace('$', ''))
                
                references.append(ref)
                
            except (ValueError, IndexError) as e:
                logger.warning(f"Could not parse cell reference from formula: {match}")
                continue
        
        return references
    
    def analyze_formula_complexity(self, formula: str) -> Dict[str, Any]:
        """Analyze the complexity of a formula"""
        if not formula or not formula.startswith('='):
            return {"complexity": "simple", "can_drill_down": False, "has_external_refs": False}
        
        # Check for external references
        has_external_refs = self._has_external_references(formula)
        
        # Basic complexity check (e.g., number of functions, references)
        num_refs = len(self.parse_formula(formula))
        num_functions = sum(1 for func in self.supported_functions if f"{func}(" in formula.upper())
        
        complexity = "simple"
        if num_refs > 5 or num_functions > 2:
            complexity = "complex"
        elif num_refs > 1 or num_functions > 0:
            complexity = "moderate"
            
        return {
            "complexity": complexity,
            "can_drill_down": not has_external_refs and num_refs > 0,
            "has_external_refs": has_external_refs
        }
    
    def get_formula_function(self, formula: str) -> Optional[str]:
        """Extract the main function from a formula"""
        if not formula or not formula.startswith('='):
            return None
            
        # Look for function patterns
        formula_upper = formula.upper()
        for func in self.supported_functions:
            if f"{func}(" in formula_upper:
                return func
                
        return None
    
    def build_dependency_tree(self, workbook_path: Path, sheet_name: str, cell_address: str, max_depth: int = 10) -> Optional[FormulaComponent]:
        """Build a dependency tree for a specific cell"""
        wb = None
        wb_data = None
        try:
            wb = openpyxl.load_workbook(workbook_path, data_only=False)
            wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
            # Normalize sheet name comparison
            target_sheet_lower = sheet_name.strip().lower()
            
            # Find the correct sheet name with case-insensitivity
            actual_sheet_name = None
            for s in wb.sheetnames:
                if s.strip().lower() == target_sheet_lower:
                    actual_sheet_name = s
                    break
            
            if not actual_sheet_name:
                logger.warning(f"Sheet '{sheet_name}' not found. Available sheets: {wb.sheetnames}")
                return None
            
            sheet_name = actual_sheet_name
            component = self._analyze_cell_recursive(
                wb, wb_data, sheet_name, cell_address, max_depth=max_depth, current_depth=0
            )
            return component
        except Exception as e:
            logger.error(f"Error building dependency tree: {e}")
            return None
        finally:
            if wb:
                try:
                    wb.close()
                except Exception:
                    pass
            if wb_data:
                try:
                    wb_data.close()
                except Exception:
                    pass
    
    def _analyze_cell_recursive(self, wb, wb_data, sheet_name: str, cell_address: str,
                              max_depth: int, current_depth: int,
                              visited_cells: Optional[Set] = None, visiting: Optional[Set] = None) -> Optional[FormulaComponent]:
        """Recursively analyze a cell and its dependencies"""
        # Initialize sets at the top level call
        if visited_cells is None:
            visited_cells = set()
        if visiting is None:
            visiting = set()
            
        # Create unique cell identifier
        sheet_part = f"{sheet_name.strip()}!" if sheet_name else ""
        cell_address = (cell_address or "A1").strip()
        cell_id = f"{sheet_part}{cell_address}"
        
        # Check for circular references in current path
        if cell_id in visiting:
            logger.warning(f"Circular reference detected in current path: {cell_id}")
            return None
            
        if cell_id in visited_cells:
            return None  # Already fully processed
            
        if current_depth >= max_depth:
            logger.warning(f"Max depth {max_depth} reached for {cell_id}")
            return None

        # Track current cell in visiting set
        visiting.add(cell_id)
        component = None
        
        try:
            # Check if the target sheet exists
            if sheet_name not in wb.sheetnames:
                logger.warning(f"Sheet '{sheet_name}' not found in workbook")
                return None

            # Get the cell with formulas
            sheet = wb[sheet_name]
            sheet_data = wb_data[sheet_name]
            
            # Parse cell address
            cell_match = self.simple_cell_pattern.match(cell_address)
            if not cell_match:
                return None
                
            col, row = cell_match.groups()
            col = col.upper()  # Normalize column letters
            row = int(row)
            
            # Get cell objects
            cell = sheet[f"{col}{row}"]
            cell_data = sheet_data[f"{col}{row}"]
            
            # Get cell value
            value = 0.0
            if cell_data.value is not None and isinstance(cell_data.value, (int, float)):
                value = float(cell_data.value)
                
            # Create component
            component = FormulaComponent(
                name=f"{sheet_name}!{cell_address}",
                cell_reference=f"{sheet_name}!{cell_address}",
                value=value,
                formula=cell.value if cell.data_type == 'f' else None,
                is_leaf_node=cell.data_type != 'f'
            )
            
            # Analyze formula dependencies
            if cell.data_type == 'f' and cell.value:
                formula = str(cell.value)
                
                if self._has_external_references(formula):
                    logger.info(f"External reference detected in {cell_id}, stopping drill-down")
                    component.is_leaf_node = True
                    return component
                    
                references = self.parse_formula(formula)
                
                for ref in references:
                    if ref.is_range:
                        range_components = self._expand_range_reference(
                            wb, wb_data, ref, max_depth, current_depth + 1, visited_cells, sheet_name
                        )
                        component.dependencies.extend(range_components)
                    else:
                        target_sheet = ref.sheet_name or sheet_name
                        target_address = f"{ref.column}{ref.row}"
                        
                        # If we're at the depth limit, create a simple component without recursing
                        if current_depth + 1 >= max_depth:
                            # Get basic cell info without recursing further
                            try:
                                sheet = wb[target_sheet]
                                sheet_data = wb_data[target_sheet]
                                
                                cell_match = self.simple_cell_pattern.match(target_address)
                                if cell_match:
                                    col, row = cell_match.groups()
                                    col = col.upper()
                                    row = int(row)
                                    
                                    cell = sheet[f"{col}{row}"]
                                    cell_data = sheet_data[f"{col}{row}"]
                                    
                                    value = 0.0
                                    if cell_data.value is not None and isinstance(cell_data.value, (int, float)):
                                        value = float(cell_data.value)
                                    
                                    dep_component = FormulaComponent(
                                        name=f"{target_sheet}!{target_address}",
                                        cell_reference=f"{target_sheet}!{target_address}",
                                        value=value,
                                        formula=str(cell.value) if cell.data_type == 'f' else None,
                                        is_leaf_node=cell.data_type != 'f'  # True leaf only if no formula
                                    )
                                    component.dependencies.append(dep_component)
                            except Exception as e:
                                logger.warning(f"Error creating shallow dependency for {target_sheet}!{target_address}: {e}")
                        else:
                            dep_component = self._analyze_cell_recursive(
                                wb, wb_data, target_sheet, target_address,
                                max_depth, current_depth + 1, visited_cells, visiting
                            )
                            if dep_component:
                                component.dependencies.append(dep_component)
            return component
            
        finally:
            # Update tracking sets
            visiting.discard(cell_id)
            if component:  # Only add to visited if we successfully processed
                visited_cells.add(cell_id)
    
    def _expand_range_reference(self, wb, wb_data, ref: CellReference,
                               max_depth: int, current_depth: int, visited_cells: Set,
                               current_sheet: str) -> List[FormulaComponent]:
        """Expand a range reference into individual cell components"""
        components = []
        
        if not ref.is_range or not ref.range_end_column or ref.range_end_row is None:
            return components
            
        try:
            # Convert column letters to numbers
            start_col = self._column_letter_to_number(ref.column)
            end_col = self._column_letter_to_number(ref.range_end_column)
            
            # Process each cell in range
            for row_num in range(ref.row, ref.range_end_row + 1):
                for col_num in range(start_col, end_col + 1):
                    col_letter = self._number_to_column_letter(col_num)
                    target_address = f"{col_letter}{row_num}"
                    target_sheet = ref.sheet_name or current_sheet
                    
                    # If we're at the depth limit, create a simple component without recursing
                    if current_depth >= max_depth:
                        # Get basic cell info without recursing further
                        try:
                            sheet = wb[target_sheet]
                            sheet_data = wb_data[target_sheet]
                            
                            cell_match = self.simple_cell_pattern.match(target_address)
                            if cell_match:
                                col, row = cell_match.groups()
                                col = col.upper()
                                row = int(row)
                                
                                cell = sheet[f"{col}{row}"]
                                cell_data = sheet_data[f"{col}{row}"]
                                
                                value = 0.0
                                if cell_data.value is not None and isinstance(cell_data.value, (int, float)):
                                    value = float(cell_data.value)
                                
                                component = FormulaComponent(
                                    name=f"{target_sheet}!{target_address}",
                                    cell_reference=f"{target_sheet}!{target_address}",
                                    value=value,
                                    formula=str(cell.value) if cell.data_type == 'f' else None,
                                    is_leaf_node=cell.data_type != 'f'  # True leaf only if no formula
                                )
                                components.append(component)
                        except Exception as e:
                            logger.warning(f"Error creating shallow range component for {target_sheet}!{target_address}: {e}")
                    else:
                        component = self._analyze_cell_recursive(
                            wb, wb_data,
                            target_sheet,
                            target_address,
                            max_depth,
                            current_depth,
                            visited_cells
                        )
                        if component:
                            components.append(component)
                        
        except Exception as e:
            logger.error(f"Error expanding range {ref}: {str(e)}")
            
        return components
    
    def _column_letter_to_number(self, column_letter: str) -> int:
        """Convert Excel column letter to number (A=1, B=2, etc.)"""
        return sum((ord(char.upper()) - 64) * 26**i for i, char in enumerate(reversed(column_letter)))
    
    def _number_to_column_letter(self, column_number: int) -> str:
        """Convert number to Excel column letter"""
        letters = []
        while column_number > 0:
            column_number, remainder = divmod(column_number - 1, 26)
            letters.append(chr(65 + remainder))
        return ''.join(reversed(letters))
    
    def _has_external_references(self, formula: str) -> bool:
        """Check if formula contains external file references"""
        return any(re.search(pattern, formula, re.IGNORECASE)
                 for pattern in [r'\[[^]]+\.xlsx?\]', r"'\[[^']+\.[^']+\.xlsx?\]"])

    def get_progressive_dependencies(self, workbook_path: Path, sheet_name: str, cell_address: str, depth: int = 1) -> Optional[List[FormulaComponent]]:
        """Get progressive dependencies for a cell up to a certain depth."""
        tree = self.build_dependency_tree(workbook_path, sheet_name, cell_address, max_depth=depth)
        return tree.dependencies if tree else None
