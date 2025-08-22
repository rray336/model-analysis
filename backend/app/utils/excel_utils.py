"""
Excel utility functions for parsing and analysis
"""
import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter, column_index_from_string
from typing import Dict, List, Tuple, Optional, Any
from pathlib import Path
import re
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

class ExcelReader:
    """Utility class for reading and analyzing Excel files"""
    
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self.workbook = None
        self.sheets_info = {}
    
    def load_workbook(self) -> openpyxl.Workbook:
        """Load the Excel workbook"""
        try:
            # Use read_only mode for better performance with large files
            self.workbook = openpyxl.load_workbook(self.file_path, read_only=True, data_only=True)
            logger.info(f"Loaded workbook: {self.file_path.name}")
            return self.workbook
        except Exception as e:
            logger.error(f"Failed to load workbook {self.file_path}: {str(e)}")
            raise
    
    def get_sheet_names(self) -> List[str]:
        """Get all sheet names in the workbook"""
        if not self.workbook:
            self.load_workbook()
        return self.workbook.sheetnames if self.workbook else []
    
    def analyze_sheet_content(self, sheet_name: str, max_rows: int = 50, max_cols: int = 20) -> Dict[str, Any]:
        """Analyze the content of a specific sheet"""
        if not self.workbook:
            self.load_workbook()
        
        if not self.workbook:
            return {}
            
        sheet = self.workbook[sheet_name]
        
        # Get basic sheet info - limit dimensions for performance
        actual_max_row = min(sheet.max_row or 0, max_rows)
        actual_max_col = min(sheet.max_column or 0, max_cols)
        
        info = {
            'name': sheet_name,
            'max_row': sheet.max_row,
            'max_column': sheet.max_column,
            'analyzed_rows': actual_max_row,
            'analyzed_columns': actual_max_col,
            'has_data': sheet.max_row and sheet.max_row > 1,
        }
        
        # Analyze sample rows to identify structure
        sample_rows = []
        for row in range(1, actual_max_row + 1):
            row_data = []
            for col in range(1, actual_max_col + 1):
                cell = sheet.cell(row=row, column=col)
                # Get coordinate manually since read_only mode might not have it
                coord = f"{get_column_letter(col)}{row}"
                row_data.append({
                    'value': cell.value,
                    'formula': None,  # Formulas not available in data_only mode
                    'data_type': 'n' if isinstance(cell.value, (int, float)) else 's' if cell.value else None,
                    'coordinate': coord
                })
            sample_rows.append(row_data)
        
        info['sample_data'] = sample_rows
        return info
    
    def sheet_to_dataframe(self, sheet_name: str, header_row: int = 0) -> pd.DataFrame:
        """Convert sheet to pandas DataFrame"""
        try:
            df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=header_row)
            return df
        except Exception as e:
            logger.warning(f"Failed to convert sheet {sheet_name} to DataFrame: {str(e)}")
            # Fallback: read without headers
            try:
                df = pd.read_excel(self.file_path, sheet_name=sheet_name, header=None)
                return df
            except Exception as e2:
                logger.error(f"Complete failure reading sheet {sheet_name}: {str(e2)}")
                return pd.DataFrame()

def get_cell_value_and_formula(workbook_path: Path, sheet_name: str, cell_address: str) -> Tuple[Optional[float], Optional[str]]:
    """Get both the calculated value and formula for a specific cell"""
    wb_formula = None
    wb_data = None
    try:
        # Load workbook with data_only=False to get formulas
        wb_formula = openpyxl.load_workbook(workbook_path, data_only=False)
        # Load workbook with data_only=True to get calculated values
        wb_data = openpyxl.load_workbook(workbook_path, data_only=True)
        
        # Find the correct sheet name with case-insensitivity
        actual_sheet_name = None
        for s in wb_formula.sheetnames:
            if s.strip().lower() == sheet_name.strip().lower():
                actual_sheet_name = s
                break
        
        if not actual_sheet_name:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return None, None
        
        sheet_formula = wb_formula[actual_sheet_name]
        sheet_data = wb_data[actual_sheet_name]
        
        # Get the cell
        cell_formula = sheet_formula[cell_address]
        cell_data = sheet_data[cell_address]
        
        # Extract value
        value = None
        if cell_data.value is not None and isinstance(cell_data.value, (int, float)):
            value = float(cell_data.value)
        
        # Extract formula
        formula = None
        if cell_formula.data_type == 'f' and cell_formula.value:
            formula = str(cell_formula.value)
        
        return value, formula
        
    except Exception as e:
        logger.error(f"Error getting cell {sheet_name}!{cell_address}: {e}")
        return None, None
    finally:
        # Ensure workbooks are always closed
        if wb_formula:
            try:
                wb_formula.close()
            except Exception:
                pass
        if wb_data:
            try:
                wb_data.close()
            except Exception:
                pass

def validate_cell_address(cell_address: str) -> bool:
    """Validate that a cell address is in the correct format (e.g., A1, B5, AC123)"""
    pattern = r'^[A-Z]+\d+$'
    return bool(re.match(pattern, cell_address.upper()))

def parse_cell_address(cell_address: str) -> Tuple[str, int]:
    """Parse a cell address into column and row components"""
    match = re.match(r'^([A-Z]+)(\d+)$', cell_address.upper())
    if not match:
        raise ValueError(f"Invalid cell address: {cell_address}")
    
    column, row = match.groups()
    return column, int(row)

def detect_financial_keywords(text: str) -> Dict[str, int]:
    """
    Detect financial keywords in text and return match scores
    """
    if not isinstance(text, str):
        return {}
    
    text_lower = text.lower()
    
    keyword_groups = {
        'income_statement': [
            'income', 'profit', 'loss', 'p&l', 'revenue', 'sales', 'ebitda', 'ebit',
            'operating', 'gross profit', 'net income', 'earnings', 'margin'
        ],
        'balance_sheet': [
            'balance', 'sheet', 'assets', 'liabilities', 'equity', 'cash', 'debt',
            'current assets', 'fixed assets', 'retained earnings', 'stockholder'
        ],
        'cash_flow': [
            'cash flow', 'operating cash', 'investing', 'financing', 'capex',
            'free cash flow', 'working capital', 'depreciation'
        ]
    }
    
    scores = {}
    for category, keywords in keyword_groups.items():
        score = sum(1 for keyword in keywords if keyword in text_lower)
        scores[category] = score
    
    return scores

def get_row_values(workbook_path: Path, sheet_name: str, row_number: int, max_columns: int = 6) -> List[Dict[str, Any]]:
    """Get values from a specific row for column selection dropdown"""
    wb = None
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return []
        
        sheet = wb[sheet_name]
        row_values = []
        
        for col in range(1, min(max_columns + 1, sheet.max_column + 1)):
            cell = sheet.cell(row=row_number, column=col)
            column_letter = get_column_letter(col)
            
            # Convert value to string for display
            display_value = ""
            if cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    display_value = str(cell.value)
                elif isinstance(cell.value, datetime):
                    display_value = cell.value.strftime("%Y-%m-%d")
                else:
                    display_value = str(cell.value)
            
            row_values.append({
                "column": column_letter,
                "value": display_value,
                "is_meaningful": len(display_value.strip()) > 0 and not display_value.isdigit()
            })
        
        return row_values
        
    except Exception as e:
        logger.error(f"Error getting row values for {sheet_name}!{row_number}: {e}")
        return []
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

def get_column_values(workbook_path: Path, sheet_name: str, column_letter: str, max_rows: int = 5) -> List[Dict[str, Any]]:
    """Get values from a specific column for row value selection dropdown"""
    wb = None
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            logger.warning(f"Sheet '{sheet_name}' not found in workbook")
            return []
        
        sheet = wb[sheet_name]
        column_values = []
        
        # Convert column letter to index
        try:
            col_index = column_index_from_string(column_letter)
        except ValueError:
            logger.error(f"Invalid column letter: {column_letter}")
            return []
        
        for row in range(1, min(max_rows + 1, sheet.max_row + 1)):
            cell = sheet.cell(row=row, column=col_index)
            
            # Convert value to string for display
            display_value = ""
            if cell.value is not None:
                if isinstance(cell.value, (int, float)):
                    display_value = str(cell.value)
                elif isinstance(cell.value, datetime):
                    display_value = cell.value.strftime("%Y-%m-%d")
                else:
                    display_value = str(cell.value)
            
            column_values.append({
                "row": str(row),
                "value": display_value,
                "is_meaningful": len(display_value.strip()) > 0 and not display_value.isdigit()
            })
        
        return column_values
        
    except Exception as e:
        logger.error(f"Error getting column values for {sheet_name}!{column_letter}: {e}")
        return []
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

def get_cell_name_from_column(workbook_path: Path, sheet_name: str, row_number: int, column_letter: str) -> Optional[str]:
    """Get the name value from a specific column of a row"""
    wb = None
    try:
        wb = openpyxl.load_workbook(workbook_path, data_only=True)
        
        if sheet_name not in wb.sheetnames:
            return None
        
        sheet = wb[sheet_name]
        cell_address = f"{column_letter}{row_number}"
        cell = sheet[cell_address]
        
        if cell.value is not None:
            if isinstance(cell.value, (int, float)):
                return str(cell.value)
            elif isinstance(cell.value, datetime):
                return cell.value.strftime("%Y-%m-%d")
            else:
                return str(cell.value).strip()
        
        return None
        
    except Exception as e:
        logger.error(f"Error getting cell name from {sheet_name}!{column_letter}{row_number}: {e}")
        return None
    finally:
        if wb:
            try:
                wb.close()
            except Exception:
                pass

def analyze_cell_relationships(sheet) -> Dict[str, List[str]]:
    """
    Analyze formula relationships in a sheet
    Returns mapping of cells to their dependencies
    """
    relationships = {}
    
    for row in sheet.iter_rows():
        for cell in row:
            if cell.data_type == 'f' and cell.value:  # Formula cell
                formula = cell.value
                # Extract cell references from formula
                cell_refs = re.findall(r'[A-Z]+\d+', formula)
                if cell_refs:
                    relationships[cell.coordinate] = cell_refs
    
    return relationships
